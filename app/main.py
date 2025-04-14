from fastapi import Body, FastAPI, HTTPException, Depends, Path, Query, status, Security
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel, EmailStr, validator, Field
import re
from passlib.context import CryptContext
from datetime import datetime, timedelta
from sqlalchemy.orm import Session
from typing import Any, List
from app.crud import calculate_course_averages, calculate_student_averages, create_grade, get_grades_by_student, update_grade, delete_grade, bulk_create_grades
from typing import Optional
from fastapi import UploadFile, File
from fastapi.responses import JSONResponse, StreamingResponse
import csv
import io
import os
from typing import Dict, List, Optional
import openpyxl
from tempfile import NamedTemporaryFile
from app.validators import GradeValidator
from app.crud import create_student, get_student, get_students, update_student, delete_student
from apscheduler.schedulers.background import BackgroundScheduler
from app.backup import create_backup

import firebase_admin
from firebase_admin import credentials, auth

from app.database import GradeHistory, init_db, SessionLocal, User as DBUser

from app.crud import get_grade_history, get_student_grade_history
from app.crud import (
    create_course, get_course, get_courses,
    add_student_to_course, remove_student_from_course,
    get_students_in_course, get_courses_for_student
)
from app.database import Course, StudentCourse

DATABASE_URL = "sqlite:///app.db"  # Update with your actual database URL

def schedule_backups():
    """Schedule daily backups."""
    scheduler = BackgroundScheduler()
    scheduler.add_job(
        func=create_backup,
        trigger="cron",
        hour=2,  # Run daily at 2 AM
        args=[DATABASE_URL],
        id="daily_backup",
        replace_existing=True,
    )
    scheduler.start()

def is_excel_file(filename: str) -> bool:
    """Check if the file is an Excel file based on extension."""
    return filename.lower().endswith(('.xlsx', '.xls'))

def parse_excel_file(content: bytes):
    """Parse Excel file content and convert to list of dictionaries."""
    # Create temp file to hold Excel data
    with NamedTemporaryFile(suffix='.xlsx', delete=False) as temp:
        temp_path = temp.name
        temp.write(content)
    
    try:
        # Load Excel workbook
        try:
            workbook = openpyxl.load_workbook(temp_path, read_only=True)
            sheet = workbook.active
            
            # Get header row
            if not sheet.max_row or sheet.max_row < 2:
                raise ValueError("Excel file is empty or missing data rows")
                
            headers = [cell.value for cell in next(sheet.rows)]
            
            # Check if all required headers are present
            required_headers = ['student_id', 'subject', 'grade']
            missing_headers = [h for h in required_headers if h not in headers]
            
            if missing_headers:
                raise ValueError(f"Missing required columns: {', '.join(missing_headers)}")
            
            # Prepare data rows
            data = []
            for row in sheet.iter_rows(min_row=2):  # Skip header row
                if all(cell.value is None for cell in row):
                    continue  # Skip empty rows
                    
                row_data = {}
                for header, cell in zip(headers, row):
                    row_data[header] = str(cell.value) if cell.value is not None else ""
                data.append(row_data)
            
            return data
            
        except openpyxl.utils.exceptions.InvalidFileException:
            raise ValueError("The file is not a valid Excel file")
            
    except Exception as e:
        raise ValueError(f"Excel parsing error: {str(e)}")
        
    finally:
        # Clean up temp file
        import os
        if os.path.exists(temp_path):
            os.unlink(temp_path)

# ----------------------------
# Firebase Admin Initialization
# ----------------------------
cred = credentials.Certificate("app/cert/serviceAccountKey.json")
firebase_admin.initialize_app(cred)

# ----------------------------
# FastAPI App and Database Setup
# ----------------------------
app = FastAPI()

# After creating the app instance
app = FastAPI()

@app.exception_handler(Exception)
async def global_exception_handler(request, exc):
    """Handle unexpected exceptions gracefully"""
    # Don't expose internal error details in production
    error_message = str(exc) if os.getenv("DEBUG") == "1" else "An unexpected error occurred"
    
    # Log the full error for debugging
    import traceback
    traceback.print_exc()
    
    return JSONResponse(
        status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        content={"detail": error_message},
    )

init_db()  # Create tables if they don't exist

# ----------------------------
# Local Security (for signup password hashing)
# ----------------------------
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
# For Firebase token extraction from headers.
firebase_scheme = HTTPBearer()

# ----------------------------
# Pydantic Models
# ----------------------------
class User(BaseModel):
    username: str
    email: str

    class Config:
        orm_mode = True  # Enables compatibility with SQLAlchemy models

class UserCreate(BaseModel):
    username: str = Field(..., min_length=3, max_length=50)
    email: EmailStr  # Validates email format
    password: str = Field(..., min_length=8)
    
    @validator('username')
    def username_alphanumeric(cls, v):
        if not re.match(r'^[a-zA-Z0-9_-]+$', v):
            raise ValueError('Username must contain only letters, numbers, underscores, and hyphens')
        return v
        
    @validator('password')
    def password_strength(cls, v):
        if not any(char.isdigit() for char in v):
            raise ValueError('Password must contain at least one digit')
        if not any(char.isalpha() for char in v):
            raise ValueError('Password must contain at least one letter')
        return v

class Token(BaseModel):
    access_token: str
    token_type: str

class GradeBase(BaseModel):
    subject: str = Field(..., min_length=1, max_length=100)
    grade: int = Field(..., ge=0, le=100)  # ge=greater or equal, le=less or equal

class GradeCreate(GradeBase):
    student_id: int = Field(..., gt=0)  # gt=greater than 0

class GradeUpdate(BaseModel):
    grade: int = Field(..., ge=0, le=100)

class GradeResponse(GradeBase):
    id: int
    student_id: int

    class Config:
        orm_mode = True
        from_attributes = True

class UserProfileUpdate(BaseModel):
    email: Optional[EmailStr] = None
    password: Optional[str] = Field(None, min_length=8)
    
    @validator('password')
    def password_strength(cls, v):
        if v is None:
            return v
        if not any(char.isdigit() for char in v):
            raise ValueError('Password must contain at least one digit')
        if not any(char.isalpha() for char in v):
            raise ValueError('Password must contain at least one letter')
        return v

class BulkUploadResponse(BaseModel):
    total_processed: int
    successful: int
    failed: int
    errors: Optional[List[str]] = None

class GradeHistoryResponse(BaseModel):
    id: int
    grade_id: int
    student_id: int
    subject: str
    old_value: Optional[int]
    new_value: Optional[int]
    action: str
    timestamp: str
    changed_by: Optional[str]
    
    class Config:
        orm_mode = True
        from_attributes = True

class PaginatedResponse(BaseModel):
    items: List[GradeHistoryResponse]
    total: int
    page: int
    pages: int
    limit: int

class CourseBase(BaseModel):
    name: str = Field(..., min_length=1, max_length=100)
    description: Optional[str] = None

class CourseCreate(CourseBase):
    teacher_id: Optional[int] = None

class CourseResponse(CourseBase):
    id: int
    teacher_id: Optional[int]
    created_at: str
    
    class Config:
        orm_mode = True
        from_attributes = True

class StudentCourseResponse(BaseModel):
    id: int
    student_id: int
    course_id: int
    joined_at: str
    added_by: Optional[str]
    
    class Config:
        orm_mode = True
        from_attributes = True

class EnrollmentResponse(BaseModel):
    message: str
    enrollment: Optional[StudentCourseResponse] = None

class StudentAverageResponse(BaseModel):
    student_id: int
    subject_averages: Dict[str, float]
    overall_average: Optional[float]
    total_grades: int

class CourseAverageResponse(BaseModel):
    course_id: int
    student_averages: List[Dict[str, Any]]
    subject_averages: Dict[str, float]
    overall_average: Optional[float]
    total_students: int

class Student(BaseModel):
    id: int
    name: str
    email: str
    date_of_birth: str

    class Config:
        orm_mode = True

# ----------------------------
# Database Dependency
# ----------------------------
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# ----------------------------
# Utility Functions (Local Password Handling)
# ----------------------------
def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

# ----------------------------
# Firebase Authentication Dependencies
# ----------------------------
def get_current_firebase_user(token: HTTPAuthorizationCredentials = Depends(firebase_scheme)) -> dict:
    """
    Verifies the Firebase ID token and returns its decoded payload.
    """
    if token is None or not token.credentials:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authentication token is missing"
        )
        
    # Test mode bypass
    # IMPORTANT: Only for development/testing, remove in production!
    if os.environ.get("TEST_MODE") == "1" or (token and token.credentials == "test-token"):
        return {
            "uid": "test-user-id",
            "email": "test@example.com",
            "roles": ["admin", "teacher"]  # Give test user necessary roles
        }
    
    
    try:
        decoded_token = auth.verify_id_token(token.credentials)
        return decoded_token
    except auth.ExpiredIdTokenError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authentication token has expired. Please log in again."
        )
    except auth.InvalidIdTokenError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid authentication token. Please log in again."
        )
    except auth.RevokedIdTokenError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authentication token has been revoked. Please log in again."
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail=f"Authentication error: {str(e)}"
        )
    
def require_roles(required_roles: List[str]):
    """
    Dependency generator that checks whether the Firebase token includes at least one
    of the required roles (e.g., "admin", "teacher", "student"). It assumes that the
    Firebase custom claims include a "roles" key.
    """
    def role_checker(user: dict = Depends(get_current_firebase_user)) -> dict:
        roles = user.get("roles", [])
        if isinstance(roles, str):
            roles = [roles]
        if not any(role in roles for role in required_roles):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Insufficient permissions"
            )
        return user
    return role_checker

@app.post("/backup", response_model=dict)
def manual_backup(db_url: str = DATABASE_URL):
    """Manually trigger a database backup."""
    try:
        backup_file = create_backup(db_url)
        return {"message": "Backup created successfully", "file": backup_file}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Backup failed: {str(e)}")
# ----------------------------
# Authentication Endpoints
# ----------------------------

@app.post("/auth/signup", response_model=User)
def signup(user: UserCreate, db: Session = Depends(get_db)):
    # Check if the user already exists locally
    existing_user = db.query(DBUser).filter(DBUser.username == user.username).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="Username already registered")
    
    # Create new user locally (storing hashed password for local reference)
    hashed_password = get_password_hash(user.password)
    new_user = DBUser(username=user.username, email=user.email, hashed_password=hashed_password)
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return new_user

# Note: In a Firebase-driven system, the client typically handles authentication.
# For demonstration, the login endpoint below simply verifies the provided Firebase ID token.
@app.post("/auth/login", response_model=Token)
def login(credentials: HTTPAuthorizationCredentials = Depends(firebase_scheme), db: Session = Depends(get_db)):
    try:
        # Verify the Firebase token sent by the client
        decoded_token = auth.verify_id_token(credentials.credentials)
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid Firebase token")
    # Optionally, you could update the local user record or sync roles.
    return {"access_token": credentials.credentials, "token_type": "Bearer"}

# Dependency to retrieve the current user from Firebase token and match with local DB
def get_current_user(token: HTTPAuthorizationCredentials = Depends(firebase_scheme), db: Session = Depends(get_db)) -> DBUser:
    try:
        decoded_token = auth.verify_id_token(token.credentials)
        username = decoded_token.get("sub")  # Using the 'sub' claim as username
        if username is None:
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token: no subject")
    except Exception:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid Firebase token")
    
    user = db.query(DBUser).filter(DBUser.username == username).first()
    if user is None:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="User not found")
    return user

# ----------------------------
# Protected Endpoints
# ----------------------------
@app.get("/users/me", response_model=User)
def read_users_me(current_user: DBUser = Depends(get_current_user)):
    return current_user

@app.get("/admin/dashboard")
def admin_dashboard(user: dict = Depends(require_roles(["admin"]))):
    return {"message": "Welcome to the admin dashboard!", "user": user}

@app.get("/teacher/portal")
def teacher_portal(user: dict = Depends(require_roles(["teacher"]))):
    return {"message": "Welcome to the teacher portal!", "user": user}

@app.get("/student/area")
def student_area(user: dict = Depends(require_roles(["student"]))):
    return {"message": "Welcome to the student area!", "user": user}

# ----------------------------
# Grades Endpoints
# ----------------------------

@app.post("/grades/", response_model=GradeResponse)
def add_grade(
    grade: GradeCreate, 
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_roles(["admin", "teacher"]))
):
    try:
        # Extract user identifier (email or UID)
        user_identifier = current_user.get("email", current_user.get("uid", "unknown"))
        
        created_grade = create_grade(
            db, 
            student_id=grade.student_id, 
            subject=grade.subject, 
            grade=grade.grade,
            changed_by=user_identifier
        )
        return GradeResponse.from_orm(created_grade)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/grades/{student_id}", response_model=List[GradeResponse])
def list_grades(student_id: int = Path(..., gt=0, description="The student ID"), db: Session = Depends(get_db)):
    grades = get_grades_by_student(db, student_id)
    return [GradeResponse.from_orm(grade) for grade in grades]


@app.put("/grades/{grade_id}", response_model=GradeResponse)
def modify_grade(
    grade_id: int = Path(..., gt=0), 
    grade_update: GradeUpdate = Body(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_roles(["admin", "teacher"]))
):
    try:
        # Extract user identifier (email or UID)
        user_identifier = current_user.get("email", current_user.get("uid", "unknown"))
        
        updated_grade = update_grade(
            db, 
            grade_id, 
            grade_update.grade,
            changed_by=user_identifier
        )
        if not updated_grade:
            raise HTTPException(status_code=404, detail="Grade not found")
        return GradeResponse.from_orm(updated_grade)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.delete("/grades/{grade_id}", response_model=GradeResponse)
def remove_grade(
    grade_id: int = Path(..., gt=0), 
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_roles(["admin", "teacher"]))
):
    # Extract user identifier (email or UID)
    user_identifier = current_user.get("email", current_user.get("uid", "unknown"))
    
    deleted_grade = delete_grade(db, grade_id, changed_by=user_identifier)
    if not deleted_grade:
        raise HTTPException(status_code=404, detail="Grade not found")
    return GradeResponse.from_orm(deleted_grade)


@app.post("/grades/upload", response_model=BulkUploadResponse)
async def upload_grades(
    file: UploadFile = File(...),
    min_grade: int = Query(0, ge=0, le=100),
    max_grade: int = Query(100, ge=0, le=100),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_roles(["admin", "teacher"]))
):
    user_identifier = current_user.get("email", current_user.get("uid", "unknown"))

    """
    Upload grades from a CSV or Excel file with comprehensive error handling.
    
    The file should have these columns:
    - student_id: The student ID (integer)
    - subject: Subject name (string)
    - grade: Grade value (integer, min_grade-max_grade)
    """
    # Check if file is empty
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    
    # Create validator with specified range
    validator = GradeValidator(min_grade=min_grade, max_grade=max_grade)
    
    # Ensure min_grade <= max_grade
    if min_grade > max_grade:
        raise HTTPException(
            status_code=400,
            detail=f"min_grade ({min_grade}) cannot be greater than max_grade ({max_grade})"
        )
    
    # Validate file size and read content
    try:
        # Reset file position to the beginning just in case
        await file.seek(0)
        
        # Check file size limit (10MB)
        MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
        content = await file.read(MAX_FILE_SIZE + 1)  # Read slightly more to check if file exceeds limit
        
        if len(content) > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=400,
                detail=f"File too large: maximum size is 10MB"
            )
            
        # Process based on file type
        try:
            # Check file type and parse accordingly
            if is_excel_file(file.filename):
                try:
                    grades_data = parse_excel_file(content)
                except ImportError:
                    raise HTTPException(
                        status_code=400,
                        detail="Excel support requires the openpyxl package. Install with: pip install openpyxl"
                    )
                except ValueError as e:
                    raise HTTPException(status_code=400, detail=str(e))
            elif file.filename.endswith('.csv'):
                # Parse as CSV
                try:
                    content_str = content.decode('utf-8')
                except UnicodeDecodeError:
                    raise HTTPException(
                        status_code=400, 
                        detail="File encoding error: Please ensure your CSV file uses UTF-8 encoding"
                    )
                    
                csv_reader = csv.DictReader(io.StringIO(content_str))
                grades_data = list(csv_reader)
                
                # Check if CSV parsing succeeded
                if not csv_reader.fieldnames:
                    raise HTTPException(
                        status_code=400,
                        detail="Invalid CSV format: could not detect column headers"
                    )
            else:
                raise HTTPException(
                    status_code=400, 
                    detail="Only CSV and Excel files are supported"
                )
            
            # Check for empty file
            if not grades_data:
                raise HTTPException(
                    status_code=400,
                    detail="File contains no data rows"
                )
                
            # Enforce maximum number of rows
            MAX_ROWS = 1000
            if len(grades_data) > MAX_ROWS:
                raise HTTPException(
                    status_code=400,
                    detail=f"Too many rows in file: maximum is {MAX_ROWS}, found {len(grades_data)}"
                )
                
            # Validate required columns
            required_columns = ['student_id', 'subject', 'grade']
            if not all(col in grades_data[0] for col in required_columns):
                missing = [col for col in required_columns if col not in grades_data[0]]
                raise HTTPException(
                    status_code=400,
                    detail=f"Missing required columns: {', '.join(missing)}"
                )
            
            # Bulk create grades with specified validator
            successful_grades, errors = bulk_create_grades(db, grades_data, validator=validator, changed_by=user_identifier)
            
            # Prepare response
            response = {
                "total_processed": len(grades_data),
                "successful": len(successful_grades),
                "failed": len(grades_data) - len(successful_grades),
            }
            
            # Add errors if any
            if errors:
                response["errors"] = errors
                
            return response
                
        except Exception as e:
            # Handle other unexpected errors during file processing
            raise HTTPException(
                status_code=500,
                detail=f"Error processing file: {str(e)}"
            )
                
    except HTTPException:
        # Re-raise HTTP exceptions directly
        raise
    except Exception as e:
        # Handle unexpected errors during file reading
        raise HTTPException(
            status_code=500,
            detail=f"Error reading file: {str(e)}"
        )

@app.get("/grades/upload/template")
async def get_grade_upload_template(
    format: str = Query("csv", pattern="^(csv|excel)$"),
    min_grade: int = Query(0, ge=0, le=100),
    max_grade: int = Query(100, ge=0, le=100),
    _: dict = Depends(require_roles(["admin", "teacher"]))
):
    # Ensure min_grade <= max_grade
    if min_grade > max_grade:
        raise HTTPException(
            status_code=400,
            detail=f"min_grade ({min_grade}) cannot be greater than max_grade ({max_grade})"
        )
    """
    Get a template file for bulk grade upload.
    
    Query parameters:
    - format: "csv" or "excel" (default: "csv")
    - min_grade: Minimum allowed grade (default: 0)
    - max_grade: Maximum allowed grade (default: 100)
    """
    # Sample data
    sample_data = [
        {"student_id": "1", "subject": f"Math (Range: {min_grade}-{max_grade})", "grade": str(min(95, max_grade))},
        {"student_id": "2", "subject": "Science", "grade": str(min(87, max_grade))},
        {"student_id": "3", "subject": "History", "grade": str(min(78, max_grade))}
    ]
    
    # If Excel format is requested
    if format.lower() == "excel":
        try:
            # Create Excel workbook and active sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            
            # Add headers
            headers = ["student_id", "subject", "grade"]
            for col_idx, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_idx, value=header)
            
            # Add sample data
            for row_idx, data_row in enumerate(sample_data, 2):
                for col_idx, header in enumerate(headers, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=data_row[header])
            
            # Save to temporary file
            with NamedTemporaryFile(delete=False, suffix='.xlsx') as temp:
                temp_path = temp.name
                workbook.save(temp_path)
                
            # Read the file and return it
            with open(temp_path, "rb") as f:
                excel_data = f.read()
                
            # Clean up the temporary file
            import os
            if os.path.exists(temp_path):
                os.unlink(temp_path)
                
            return StreamingResponse(
                io.BytesIO(excel_data),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=grade_upload_template.xlsx"}
            )
        except ImportError:
            raise HTTPException(
                status_code=400,
                detail="Excel template requires the openpyxl package. Install with: pip install openpyxl"
            )
    
    # Default to CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(['student_id', 'subject', 'grade'])
    
    # Write sample data
    for row in sample_data:
        writer.writerow([row['student_id'], row['subject'], row['grade']])
    
    # Reset buffer position
    output.seek(0)
    
    # Return CSV file
    return StreamingResponse(
        output, 
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=grade_upload_template.csv"}
    )

@app.get("/grades/{grade_id}/history", response_model=List[GradeHistoryResponse])
def get_history_for_grade(
    grade_id: int = Path(..., gt=0), 
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_roles(["admin", "teacher"]))
):
    """Get the history/audit log for a specific grade."""
    history = get_grade_history(db, grade_id)
    return [GradeHistoryResponse.from_orm(entry) for entry in history]

@app.get("/students/{student_id}/grades/history", response_model=List[GradeHistoryResponse])
def get_history_for_student(
    student_id: int = Path(..., gt=0),
    subject: Optional[str] = None,
    start_date: Optional[str] = Query(None, description="Filter by start date (ISO format)"),
    end_date: Optional[str] = Query(None, description="Filter by end date (ISO format)"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_roles(["admin", "teacher"]))
):
    """Get the grade history for a specific student with optional filters."""
    history = get_student_grade_history(db, student_id, subject, start_date, end_date)
    return [GradeHistoryResponse.from_orm(entry) for entry in history]

@app.get("/admin/grades/history", response_model=PaginatedResponse)
def get_all_grade_history_endpoint(  # ← Renamed function to avoid conflict
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    action: Optional[str] = Query(None),
    limit: int = Query(25, ge=1, le=100),
    page: int = Query(1, ge=1),  # Use page instead of offset
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_roles(["admin"]))
):
    """Admin endpoint to view all grade history/audit logs with filtering."""
    # Calculate offset from page
    offset = (page - 1) * limit
    
    # Get count for total
    query = db.query(GradeHistory)
    if start_date:
        query = query.filter(GradeHistory.timestamp >= start_date)
    if end_date:
        query = query.filter(GradeHistory.timestamp <= end_date)
    if action:
        query = query.filter(GradeHistory.action == action)
    
    total = query.count() or 0  # Ensure total is at least 0, not None
    
    # Get paginated results - use imported function from crud.py
    from app.crud import get_all_grade_history as crud_get_all_grade_history  # Import with alias
    history = crud_get_all_grade_history(db, start_date, end_date, action, limit, offset)
    history_responses = [GradeHistoryResponse.from_orm(entry) for entry in history]
    
    # Calculate total pages (safely)
    total_pages = (total + limit - 1) // limit if total > 0 else 1
    
    return {
        "items": history_responses,
        "total": total,
        "page": page,
        "pages": total_pages,
        "limit": limit
    }

# ----------------------------
# User Profile Endpoints
# ----------------------------

# Update profile endpoint
@app.put("/users/me", response_model=User)
def update_profile(
    profile_update: UserProfileUpdate,
    db: Session = Depends(get_db),
    current_user: DBUser = Depends(get_current_user)
):
    """Update the current user's profile."""
    # Check if email is being updated and if it's already taken
    if profile_update.email and profile_update.email != current_user.email:
        existing_user = db.query(DBUser).filter(DBUser.email == profile_update.email).first()
        if existing_user:
            raise HTTPException(status_code=400, detail="Email already registered")
        current_user.email = profile_update.email
    
    # Update password if provided
    if profile_update.password:
        current_user.hashed_password = get_password_hash(profile_update.password)
    
    db.commit()
    db.refresh(current_user)
    return current_user

# Admin endpoint to get any user's profile
@app.get("/users/{user_id}", response_model=User)
def get_user_profile(
    user_id: int = Path(..., gt=0, description="The user ID"), 
    db: Session = Depends(get_db),
    _: dict = Depends(require_roles(["admin"]))
):
    """Get a user's profile by ID (admin only)."""
    user = db.query(DBUser).filter(DBUser.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user

# Admin endpoint to update any user's profile
@app.put("/users/{user_id}", response_model=User)
def update_user_profile(
    user_id: int = Path(..., gt=0, description="The user ID"),
    profile_update: UserProfileUpdate = Body(...),
    db: Session = Depends(get_db),
    _: dict = Depends(require_roles(["admin"]))
):
    """Update any user's profile (admin only)."""
    user = db.query(DBUser).filter(DBUser.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Check if email is being updated and if it's already taken
    if profile_update.email and profile_update.email != user.email:
        existing_user = db.query(DBUser).filter(DBUser.email == profile_update.email).first()
        if existing_user:
            raise HTTPException(status_code=400, detail="Email already registered")
        user.email = profile_update.email
    
    # Update password if provided
    if profile_update.password:
        user.hashed_password = get_password_hash(profile_update.password)
    
    db.commit()
    db.refresh(user)
    return user

# ----------------------------
# Course Management Endpoints
# ----------------------------

@app.post("/courses/", response_model=CourseResponse)
def create_new_course(
    course: CourseCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_roles(["admin", "teacher"]))
):
    """Create a new course (admin and teachers only)"""
    # If a teacher is creating the course, automatically assign themselves
    teacher_id = None
    if "teacher" in current_user.get("roles", []) and "admin" not in current_user.get("roles", []):
        # In a real app, you'd get the teacher's database ID from their Firebase ID
        # For now, we use a placeholder approach
        teacher_id = course.teacher_id  # This would be retrieved from the user record
    
    return create_course(
        db, 
        name=course.name, 
        description=course.description,
        teacher_id=teacher_id or course.teacher_id
    )

@app.get("/courses/", response_model=List[CourseResponse])
def list_courses(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_roles(["admin", "teacher", "student"]))
):
    """List all available courses"""
    return get_courses(db, skip=skip, limit=limit)

@app.get("/courses/{course_id}", response_model=CourseResponse)
def get_course_by_id(
    course_id: int = Path(..., gt=0),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_roles(["admin", "teacher", "student"]))
):
    """Get a specific course by ID"""
    course = get_course(db, course_id)
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
    return course

# ----------------------------
# Enrollment Management
# ----------------------------

@app.post("/courses/{course_id}/students/{student_id}", response_model=EnrollmentResponse)
def add_student_endpoint(
    course_id: int = Path(..., gt=0),
    student_id: int = Path(..., gt=0),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_roles(["admin", "teacher"]))
):
    """Add a student to a course (admin and teachers only)"""
    try:
        # Get user info for audit
        user_identifier = current_user.get("email", current_user.get("uid", "unknown"))
        
        enrollment = add_student_to_course(
            db, 
            student_id=student_id, 
            course_id=course_id,
            added_by=user_identifier
        )
        return {
            "message": f"Student {student_id} added to course {course_id}",
            "enrollment": enrollment
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.delete("/courses/{course_id}/students/{student_id}", response_model=dict)
def remove_student_endpoint(
    course_id: int = Path(..., gt=0),
    student_id: int = Path(..., gt=0),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_roles(["admin", "teacher"]))
):
    """Remove a student from a course (admin and teachers only)"""
    try:
        success = remove_student_from_course(db, student_id=student_id, course_id=course_id)
        if success:
            return {"message": f"Student {student_id} removed from course {course_id}"}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/courses/{course_id}/students", response_model=List[int])
def list_students_in_course(
    course_id: int = Path(..., gt=0),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_roles(["admin", "teacher"]))
):
    """Get all student IDs enrolled in a course (admin and teachers only)"""
    return get_students_in_course(db, course_id)

@app.get("/students/{student_id}/courses", response_model=List[CourseResponse])
def list_student_courses(
    student_id: int = Path(..., gt=0),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_roles(["admin", "teacher", "student"]))
):
    """Get all courses for a student"""
    # If student is requesting, only show their own courses
    if "student" in current_user.get("roles", []) and not any(role in ["admin", "teacher"] for role in current_user.get("roles", [])):
        if str(current_user.get("uid")) != str(student_id):
            raise HTTPException(status_code=403, detail="You can only view your own courses")
        
    return get_courses_for_student(db, student_id)

# Batch enrollment endpoint - useful for adding multiple students at once
@app.post("/courses/{course_id}/students", response_model=dict)
def batch_add_students(
    course_id: int = Path(..., gt=0),
    student_ids: List[int] = Body(..., embed=True),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_roles(["admin", "teacher"]))
):
    """Add multiple students to a course at once"""
    user_identifier = current_user.get("email", current_user.get("uid", "unknown"))
    
    successful = []
    failed = []
    
    for student_id in student_ids:
        try:
            add_student_to_course(db, student_id, course_id, added_by=user_identifier)
            successful.append(student_id)
        except ValueError as e:
            failed.append({"student_id": student_id, "reason": str(e)})
    
    return {
        "message": f"Added {len(successful)} students to course {course_id}",
        "successful": successful,
        "failed": failed
    }

# ----------------------------
# Grade Statistics Endpoints
# ----------------------------

@app.get("/students/{student_id}/averages", response_model=StudentAverageResponse)
def get_student_averages(
    student_id: int = Path(..., gt=0),
    course_id: Optional[int] = Query(None, description="Filter averages by course ID"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_roles(["admin", "teacher", "student"]))
):
    """
    Get a student's grade averages by subject and overall.
    
    Optionally filter by course if specified.
    """
    # If student is requesting, only show their own averages
    if "student" in current_user.get("roles", []) and not any(role in ["admin", "teacher"] for role in current_user.get("roles", [])):
        if str(current_user.get("uid")) != str(student_id):
            raise HTTPException(status_code=403, detail="You can only view your own grade averages")
    
    # Calculate and return the averages
    return calculate_student_averages(db, student_id, course_id)

@app.get("/courses/{course_id}/averages", response_model=CourseAverageResponse)
def get_course_averages(
    course_id: int = Path(..., gt=0),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_roles(["admin", "teacher"]))
):
    """
    Get average grades for all students in a course.
    
    Returns:
    - Per-student averages
    - Per-subject averages across the course
    - Overall course average
    """
    # Check if course exists
    course = get_course(db, course_id)
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
    
    # Calculate and return the averages
    return calculate_course_averages(db, course_id)

# ----------------------------
# Student Management Endpoints
# ----------------------------

@app.post("/students", response_model=Student)
def create_student_endpoint(name: str, email: str, date_of_birth: str, db: Session = Depends(get_db)):
    return create_student(db, name, email, date_of_birth)

@app.get("/students/{student_id}", response_model=Student)
def get_student_endpoint(student_id: int, db: Session = Depends(get_db)):
    student = get_student(db, student_id)
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    return student

@app.get("/students", response_model=List[Student])
def get_students_endpoint(skip: int = 0, limit: int = 10, db: Session = Depends(get_db)):
    return get_students(db, skip, limit)

@app.put("/students/{student_id}", response_model=Student)
def update_student_endpoint(student_id: int, name: str = None, email: str = None, date_of_birth: str = None, db: Session = Depends(get_db)):
    student = update_student(db, student_id, name, email, date_of_birth)
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    return student

@app.delete("/students/{student_id}")
def delete_student_endpoint(student_id: int, db: Session = Depends(get_db)):
    student = delete_student(db, student_id)
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    return {"detail": "Student deleted"}

# ----------------------------
# HTTPS Entry Point
# ----------------------------
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "app.main:app",
        host="0.0.0.0",
        port=443,
        ssl_keyfile="app/cert/key.pem",   # Update with the path to your private key file
        ssl_certfile="app/cert/cert.pem",   # Update with the path to your certificate file
        reload=True
    )